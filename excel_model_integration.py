import pandas as pd
import openpyxl

# Step 1: Load the input Excel file
input_excel_path = r"C:\Users\HridayChhabria\AppData\Local\Microsoft\WindowsApps\Scripts\Engine\rew_ip.xlsx"
excel_path = r"C:\Users\HridayChhabria\AppData\Local\Microsoft\WindowsApps\Scripts\Engine\test.xlsx"

try:
    # Load the input Excel file into a DataFrame ('Sheet1')
    data = pd.read_excel(input_excel_path, sheet_name='Sheet1')

    # Load existing workbook
    wb = openpyxl.load_workbook(excel_path)
    if 'SQL' in wb.sheetnames:
        ws = wb['SQL']
        
        # Clear existing values in 'SQL' sheet
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.value = None

        # Copy data from the DataFrame to the 'SQL' sheet
        for r_idx, row in data.iterrows():
            for c_idx, value in enumerate(row):
                ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)
    else:
        print("Sheet 'SQL' not found in the workbook.")
    
    # Ensure at least one sheet is visible
    for sheet in wb.worksheets:
        sheet.sheet_state = 'visible'
    
    # Save the updated workbook
    wb.save(excel_path)

    print("Excel updated successfully!")

except Exception as e:
    print(f"An error occurred: {e}")
